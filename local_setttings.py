#!/bin/env python
# Cedazo - Morph, Meld and Merge data - Copyright © 2023 Iwan van der Kleijn - See LICENSE.txt for conditions
#from openpyxl import Workbook
import openpyxl
from data import change_column_AdditionalNotes,  change_column_TeamRequestComment1, change_column_criticality, change_column_criticality2, change_column_fill, copy_format_column, define_criticality, delete_rows_LeadPracticeArea, delete_rows_TeamRequestStatu, find_column_Team, find_column_lead, format_colum_to_number, format_condition_iter2, insert_column, remove_rows, subtract_two_column, tick_red_cell, unmerge_cells
#change_column_AdditionalNotes, change_column_TeamRequestComment1, delete_rows_LeadPracticeArea, delete_rows_TeamRequestStatu, find_column_Team, find_column_lead,
from miargparse import parser
from openpyxl.styles import PatternFill 

args = parser.parse_args()
print(args)

wb = openpyxl.load_workbook(args.ruta_input)
ws = wb['Retain Report']

#Quita el formato de celdas compartidas
unmerge_cells(ws)

#Elimina las filas de la (1 a la 11)
remove_rows(ws,1,11)

#Cambia a todas la celdas les pone color blanco (quita el amarillo)
change_column_fill(ws)

#Metodo que da formato a la columna que se ha creado según criterio  
color_yellow = PatternFill(start_color="FFFFFF99", end_color="FFFFFF99", fill_type = "solid")
format_condition_iter2(ws, colNr= 9, condition="CRITICA", color=color_yellow)

# Metodo que inserta una columna nueva detrás de la columna H poniendo la cabecera y copiando el formato de la columna H
insert_column(ws, colNr=9, headerRow=1, headerVal='FTES. Pdtes.')

#Metodo que da formato a la columna que se ha creado 
copy_format_column(ws, colNr= 9)

#Cambiar el formato de una columna de "." a ","
format_colum_to_number(ws, column_id=9)

#Sumamos las columnas G y H 
subtract_two_column(ws)

#Poner la columna en rojo si no se ha podido restar (G Y H)
color_blue = PatternFill(start_color="99EFEB", end_color="99EFEB", fill_type = "solid")
tick_red_cell(ws, color=color_blue, colNr= 9)

#Busca la columna que queremos tratar para borrar las filas de la columna R
find_column_lead(ws)

#Borrar Filas columna R -- Eliminar las que no sean CCA_SCE_ES
delete_rows_LeadPracticeArea(ws, find_column_lead(ws))

#Buscar la columna que queremos tratar para borrar las filas de la columna C
find_column_Team(ws)

#Borrar filas columna C -- Eliminar las filas que sean igual a Draf
delete_rows_TeamRequestStatu(ws,find_column_Team(ws))

#Busca el nombre de la cabecera F y lo cambia por CRITICIDAD 
change_column_AdditionalNotes(ws)

#Busca el nombre de la cabecera O y lo cambia por CLIENTE
change_column_TeamRequestComment1(ws)

#Busca si la columna nueva 'FTES. Pdtes = 0, poner en la columna CRITICIDAD el valor ‘CUBIERTA’
color_green = PatternFill(start_color="7BF97B", end_color="7BF97B", fill_type = "solid")
change_column_criticality(ws, colNr=9, color=color_green)

#Si en la columna FTES. Pdtes. Es distinto de 0 y IsPositionCritical es igual a "Yes" pinta se pone el la columna f(CRITICIDAD) a CRITICO Y SE PINTA DE ROJO  
color_red = PatternFill(start_color="FF0000", end_color="FF0000", fill_type = "solid")
change_column_criticality2(ws, colNr_i=9, colNr_e=5, color=color_red)

#Si en la columna FTES. Pdtes. Es distinto de 0 y IsPositionCritical es igual a "No" pinta se pone el la columna f(CRITICIDAD) a UREGENTE Y SE PINTA DE AMARILLO 
color_yellow = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type = "solid")
color_blue = PatternFill(start_color="114AFF", end_color="114AFF", fill_type = "solid")
define_criticality(ws, color1=color_yellow, color2=color_blue)







#-------------------------------------------------------------
#intermedio = '.\\downloads\\out-mi.xlsx'

#wb.save(intermedio)
#wb = openpyxl.load_workbook(intermedio)
#ws = wb['Retain Report']

#Busca en la columna F1 la palabra CRITICA (May, Min, o dentro de una frase) y marca la columna C en el color morado, pero solo se puede quitar el color quitando la formula desde el Excel de formato condicional
#color_rose=PatternFill(start_color="6124BA", end_color="6124BA", fill_type = "solid")
#ws.conditional_formatting.add('C1:C2000', FormulaRule(formula=['SEARCH("*CRITICA*",F1)'], stopIfTrue=False, fill=color_rose))

#-------------------------------------------------------------

#color_white = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type = "solid")
#color_white = PatternFill(start_color="6600CC", end_color="6600CC", fill_type = "solid")
#change_column_criticality4(ws, colNr_i=9, colNr_e=5, color=color_white)

# Save the workbook to the output file
wb.save(args.ruta_output)


