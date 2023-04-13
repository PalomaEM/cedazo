#!/bin/env python
# Cedazo - Morph, Meld and Merge data - Copyright © 2023 Iwan van der Kleijn - See LICENSE.txt for conditions
#from openpyxl import Workbook
import openpyxl

from data import change_colour, change_column_AdditionalNotes, change_column_TeamRequestComment1, delete_rows_LeadPracticeArea, delete_rows_TeamRequestStatu, find_column_Team, find_column_lead, format_column, format_condition, insert_column, remove, unmerge_cells
from miargparse import parser
from openpyxl.styles import PatternFill 
#Damos la localización del fichero de entrada
#ruta_input = "C:\\Users\\lgordoga\\L99PYTHON\\SPRINT_0\\in_corto.xlsx"
#Damos la localización del fichero de salida
#ruta_output = "C:\\Users\\lgordoga\\L99PYTHON\\SPRINT_0\\out.xlsx"
# Creamos objeto wb (libro) de tipo workbook y lo cargamos con lo del excel
args = parser.parse_args()
#ruta_input = args.ruta_input
print(args)

wb = openpyxl.load_workbook(args.ruta_input)
# Creamos objeto ws (hoja), siendo la hoja activa
#ws = wb.active 
ws = wb['Retain Report']
#print('la celda A1 es:' ,ws['A1'].value)

#Metodo que desmergea las celdas de las filas a eliminar
unmerge_cells(ws)

# Metodo que sirve para borrar todas las filas de 1 a 11
#for row in ws: 
remove(ws,1,11)

# Metodo que elimina el color amarillo de todas la celdas que sean amarillas 
change_colour(ws)

# Metodo que inserta una columna nueva detrás de la columna H poniendo la cabecera y copiando el formato de la columna H
insert_column(ws, colNr=9, headerRow=1, headerVal='FTES. Pdtes.')

#Rellanamos la columna nueva FTES. Pdtes
#insert_column_pendingFTES(ws)

#Metodo que da formato a la columna que se ha creado
format_column(ws, colNr= 9)

#Metodo que da formato a la columna que se ha creado según criterio  
color_yellow = PatternFill(start_color="FFFFFF99", end_color="FFFFFF99", fill_type = "solid")
format_condition(ws, colNr= 9, condition="CRITICA", color=color_yellow)

#Borrar Filas columna R -- Eliminar las que no sean CCA_SCE_ES
delete_rows_LeadPracticeArea(ws, find_column_lead(ws))

#Borrar filas columna C -- Eliminar las filas que sean igual a Draf
delete_rows_TeamRequestStatu(ws,find_column_Team(ws))

#Busca el nombre de la cabecera F y lo cambia por CRITICIDAD 
change_column_AdditionalNotes(ws)

#Busca si la columna nueva 'FTES. Pdtes = 0, poner en la columna CRITICIDAD el valor ‘CUBIERTA’
#change_column_criticality(ws, find_column_criticality)

#Busca el nombre de la cabecera O y lo cambia por CLIENTE
change_column_TeamRequestComment1(ws)

# Save the workbook to the output file
wb.save(args.ruta_output)


