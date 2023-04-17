from copy import copy
import numbers
from types import CellType
import openpyxl
from openpyxl.styles import PatternFill 


def remove_empty(ws): 
    '''Metodo que sirve para borrar todas las filas que estan vacias (no se utiliza de momento porque no sirve para borrar
       lad filas que hay que eliminar pero tienen contenido)'''
    filas = ws.max_row
    for i in range(filas, 0, -1):
        celdas_vacias = all([cell.value is None for cell in ws[i]])
        if celdas_vacias:
            ws.delete_rows(i, 1)

def unmerge_cells(ws):
    '''Metodo que desmergea las celdas de las filas a eliminar'''

    for merge in list(ws.merged_cells):
        
        ws.unmerge_cells(range_string=str(merge))

def remove_rows(ws,ini,end): 
    '''Metodo que sirve para borrar todas las filas de 1 a 11'''
    ws.delete_rows(ini,end)


def change_column_fill(ws): 
    '''Metodo que a todas la celdas les pone color blanco (quita el amarillo) '''
    max_row = ws.max_row
    max_column = ws.max_column
    
    for i_row in range(1, max_row + 1):
       for i_column in range(1, max_column + 1):
           cell = ws.cell(row = i_row, column = i_column)
           cell.fill = PatternFill(start_color="FFFFFFFF", end_color="FFFFFFFF", fill_type = "solid")

def copyStyle(newCell, cell): 
    '''Metodo que copia el formato de una celda a una nueva '''
    if cell.has_style: 
        newCell.style = copy(cell.style) 
        newCell.font = copy(cell.font) 
        newCell.border = copy(cell.border) 
        newCell.fill = copy(cell.fill) 
        newCell.number_format = copy(cell.number_format) 
        newCell.protection = copy(cell.protection) 
        newCell.alignment = copy(cell.alignment)
  
def insert_column(ws, colNr ,headerRow , headerVal):
    '''Metodo que inserta una columna nueva detrás de la columna H poniendo la cabecera'''
    ws.insert_cols(colNr)

    ws.cell(row=headerRow, column=colNr).value = headerVal

def subtract_two_column(ws):
    """Obtiene el resultado de la resta entre la columna G Y H"""
    min, sub = ['G', 'H']
    for row in range(ws.max_row, 1, -1):
        result_cell = 'I{}'.format(row)
        minuend = ws[min + str(row)].value
        subtrahend = ws[sub + str(row)].value
    
        if minuend and subtrahend:
            try:
                ws[result_cell] =  int(minuend) - int(subtrahend)
            except:   
                print("las celdas contienen valores erroneos y no se pueden restar, se marcan en rojo")



def tick_red_cell(ws, color, colNr):
    """Marca en rojo las celdas que no tienen valor"""
    max_row = ws.max_row
    max_column = ws.max_column
    

    for i_row in range(1, max_row + 1):
       for i_column in range(1, max_column + 1):
        cell_new = ws.cell(row=i_row, column= i_column)
        if cell_new.value == None and colNr == i_column:
           cell_new.fill = color

        
def copy_format_column(ws, colNr):
    '''Metodo que da formato a la columna que se ha creado usando método range'''
    max_row = ws.max_row

    for i_row in range(1, max_row + 1):
        cell_new = ws.cell(row=i_row, column= colNr)
        cell_origin = ws.cell(row=i_row, column=colNr-1)
        copyStyle(cell_new, cell_origin)

def format_column_iter(ws, colNr):  
    '''Metodo que da formato a la columna que se ha creado usando método iter_cols'''

    for col in ws.iter_cols(min_row=1, min_col=colNr-1, max_col=colNr-1): 
            
            for cell_origin in col:
                cell_new = ws.cell(row=cell_origin.row, column=colNr)
                copyStyle(cell_new, cell_origin)
        

"""def format_condition(ws, colNr, condition, color):     
    '''Metodo que da formato a la columna que se ha creado según criterio'''   

    max_row = ws.max_row    
    for i_row in range(1, max_row + 1):   
        cell_new = ws.cell(row=i_row, column= colNr)
        cell_condition = ws.cell(row=i_row, column= colNr-3)
        if cell_condition.value == condition:
           cell_new.fill = color"""

"""def format_condition_iter(ws, colNr, condition, color):     
    '''Metodo que da formato a la columna que se ha creado según criterio utilizando iter_cols y enumerate'''  
    
    for col in ws.iter_cols(min_row=1, min_col=colNr-3, max_col=colNr-3, values_only=True): 
       
        for i, value in enumerate(col):
        
           if value == condition:
              
              cell_coor = ws.cell(row=i+1,column= colNr).coordinate
              cell = ws[cell_coor]
              cell.fill = color"""    

def format_condition_iter2(ws, colNr, condition, color):     
    '''Metodo que da formato a la columna que se ha creado según criterio utilizando iter_cols y cell'''  

    for col_condition in ws.iter_cols(min_row=1, min_col=colNr-3, max_col=colNr-3): 

        for cell_condition in col_condition:  
           if cell_condition.value == condition:
            if cell_condition.value == condition:

              cell_coor = ws.cell(row=cell_condition.row, column=colNr).coordinate
              #print('coordenada: ', cell_coor)
              #Se utiliza la coordenada de la celda (cell_coor) para obtener un objeto cell de la hoja de cálculo (ws) 
              #correspondiente a la celda en la fila y columna especificadas.
              cell = ws[cell_coor]
              #se cambia el color al nuevo objeto creado
              cell.fill = color                    

def find_column_lead(ws):
    """Metodo que busca la columna que queremos tratar"""
    for row in ws.iter_rows(): 
        for cell in row: 
            if cell.value == "LeadPracticeArea": 
                return cell.column
                
                

def delete_rows_LeadPracticeArea(ws, idcolum):
    """Una vez que hemos encontrado en el metodo find_column_lead, evaluamos recorremos las celdas evaluando el contenido """
    for column in ws.iter_cols(): 
        for cell in column: 
            if cell.column == idcolum and cell.value != 'CCA_SCE_ES' and cell.value != 'LeadPracticeArea': 
                ws.delete_rows(cell.row) 
		      

    
def find_column_Team(ws):
    """Metodo que busca la columna que queremos tratar"""
    for row in ws.iter_rows(): 
        for cell in row: 
            if cell.value == "TeamRequestStatus": 
                return cell.column
            

def delete_rows_TeamRequestStatu(ws, idcolum):
    """Una vez que hemos encontrado en el metodo find_column_lead, evaluamos recorremos las celdas evaluando el contenido """
    for column in ws.iter_cols(): 
        for cell in column: 
            if cell.column == idcolum and cell.value == 'Draft' and cell.value != 'TeamRequestStatus': 
                ws.delete_rows(cell.row) 
		      

def change_column_AdditionalNotes(ws):
    for row in ws.iter_rows(): 
        for cell in row: 
            if cell.value == "Additional Notes": 
               cell.value = "CRITICIDAD"
            


def change_column_TeamRequestComment1(ws):
    for row in ws.iter_rows():
        for cell in row:
            if cell.value == "Team Request Comment 1":
                cell.value = "CLIENTE"




